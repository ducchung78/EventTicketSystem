import json, os
from datetime import datetime

# =====================================
# ĐỌC FILE KẾT QUẢ JSON
# =====================================
with open("ket_qua_test.json", encoding="utf-8") as f:
    data = json.load(f)

results = data["chi_tiet"]
thoi_gian = data["thoi_gian"]
tong_pass = data["tong_pass"]
tong_fail = data["tong_fail"]

# =====================================
# XUẤT HTML
# =====================================
def xuat_html():
    rows = ""
    stt = 1
    for r in results:
        mau = "#d4edda" if r["status"] == "PASS" else "#f8d7da"
        mau_text = "#155724" if r["status"] == "PASS" else "#721c24"
        badge = f'<span style="background:#28a745;color:#fff;padding:2px 10px;border-radius:12px;font-weight:bold;">PASS</span>' if r["status"] == "PASS" else f'<span style="background:#dc3545;color:#fff;padding:2px 10px;border-radius:12px;font-weight:bold;">FAIL</span>'
        rows += f"""
        <tr style="background:{mau};">
            <td style="text-align:center;font-weight:bold;">{stt}</td>
            <td style="text-align:center;font-weight:bold;">{r['id']}</td>
            <td>{r['name']}</td>
            <td style="text-align:center;">{badge}</td>
            <td>{r.get('note','')}</td>
        </tr>"""
        stt += 1

    ty_le = f"{tong_pass/(tong_pass+tong_fail)*100:.1f}%"

    html = f"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<title>Kết Quả Kiểm Thử - TicketHub</title>
<style>
  body {{ font-family: Arial, sans-serif; margin: 30px; background: #f5f5f5; }}
  .header {{ background: linear-gradient(135deg, #1e3a5f, #2d6a9f); color: white; padding: 24px 30px; border-radius: 10px; margin-bottom: 24px; }}
  .header h1 {{ margin: 0 0 6px; font-size: 22px; }}
  .header p {{ margin: 0; font-size: 13px; opacity: 0.85; }}
  .summary {{ display: flex; gap: 16px; margin-bottom: 24px; }}
  .card {{ flex: 1; background: white; border-radius: 10px; padding: 18px 24px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,.08); }}
  .card .num {{ font-size: 36px; font-weight: bold; margin-bottom: 4px; }}
  .card .lbl {{ font-size: 13px; color: #666; }}
  .pass .num {{ color: #28a745; }}
  .fail .num {{ color: #dc3545; }}
  .rate .num {{ color: #007bff; }}
  table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,.08); }}
  th {{ background: #1e3a5f; color: white; padding: 12px 14px; text-align: left; font-size: 13px; }}
  td {{ padding: 10px 14px; font-size: 13px; border-bottom: 1px solid #eee; }}
  tr:last-child td {{ border-bottom: none; }}
  .section-title {{ font-size: 15px; font-weight: bold; color: #1e3a5f; margin: 24px 0 10px; padding-left: 10px; border-left: 4px solid #2d6a9f; }}
</style>
</head>
<body>
<div class="header">
  <h1>📋 Báo Cáo Kết Quả Kiểm Thử - TicketHub</h1>
  <p>Thời gian thực thi: {thoi_gian} &nbsp;|&nbsp; Công cụ: Selenium WebDriver &nbsp;|&nbsp; URL: http://localhost:8080</p>
</div>

<div class="summary">
  <div class="card pass"><div class="num">{tong_pass}</div><div class="lbl">✅ PASS</div></div>
  <div class="card fail"><div class="num">{tong_fail}</div><div class="lbl">❌ FAIL</div></div>
  <div class="card"><div class="num">{tong_pass+tong_fail}</div><div class="lbl">📊 Tổng TC</div></div>
  <div class="card rate"><div class="num">{ty_le}</div><div class="lbl">🎯 Tỷ lệ PASS</div></div>
</div>

<div class="section-title">Chi Tiết Kết Quả Kiểm Thử</div>
<table>
  <thead>
    <tr>
      <th style="width:40px">STT</th>
      <th style="width:80px">Mã TC</th>
      <th>Mô Tả Test Case</th>
      <th style="width:90px;text-align:center;">Kết Quả</th>
      <th>Ghi Chú</th>
    </tr>
  </thead>
  <tbody>
    {rows}
  </tbody>
</table>

<p style="margin-top:20px;font-size:12px;color:#999;text-align:center;">
  Báo cáo được tạo tự động bởi Selenium WebDriver &nbsp;|&nbsp; {thoi_gian}
</p>
</body>
</html>"""

    with open("ket_qua_test.html", "w", encoding="utf-8") as f:
        f.write(html)
    print("✅ Đã xuất: ket_qua_test.html")

# =====================================
# XUẤT EXCEL
# =====================================
def xuat_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        os.system("pip install openpyxl -q")
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Ket Qua Kiem Thu"

    # Màu sắc
    BLUE_DARK = "1E3A5F"
    GREEN = "28A745"
    RED = "DC3545"
    GREEN_LIGHT = "D4EDDA"
    RED_LIGHT = "F8D7DA"
    WHITE = "FFFFFF"
    GRAY = "F2F2F2"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # TIÊU ĐỀ
    ws.merge_cells("A1:E1")
    ws["A1"] = "BÁO CÁO KẾT QUẢ KIỂM THỬ - TICKETHUB"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE_DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Thời gian: {thoi_gian}  |  Công cụ: Selenium WebDriver  |  URL: http://localhost:8080"
    ws["A2"].font = Font(size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # SUMMARY
    ws.row_dimensions[3].height = 8
    headers_sum = ["PASS", "FAIL", "TỔNG", "TỶ LỆ PASS"]
    values_sum = [tong_pass, tong_fail, tong_pass+tong_fail, f"{tong_pass/(tong_pass+tong_fail)*100:.1f}%"]
    colors_sum = [GREEN, RED, BLUE_DARK, "0070C0"]

    for i, (h, v, c) in enumerate(zip(headers_sum, values_sum, colors_sum)):
        col = i + 1
        ws.cell(4, col).value = h
        ws.cell(4, col).font = Font(bold=True, color=WHITE, size=11)
        ws.cell(4, col).fill = PatternFill("solid", fgColor=c)
        ws.cell(4, col).alignment = Alignment(horizontal="center")
        ws.cell(4, col).border = border

        ws.cell(5, col).value = v
        ws.cell(5, col).font = Font(bold=True, size=16, color=c)
        ws.cell(5, col).alignment = Alignment(horizontal="center")
        ws.cell(5, col).border = border
        ws.row_dimensions[5].height = 30

    ws.row_dimensions[6].height = 8

    # HEADER BẢNG
    headers = ["STT", "Mã TC", "Mô Tả Test Case", "Kết Quả", "Ghi Chú"]
    widths = [6, 10, 45, 12, 35]
    for i, (h, w) in enumerate(zip(headers, widths)):
        col = i + 1
        cell = ws.cell(7, col)
        cell.value = h
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=BLUE_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[7].height = 22

    # DỮ LIỆU
    for idx, r in enumerate(results):
        row = 8 + idx
        is_pass = r["status"] == "PASS"
        bg = GREEN_LIGHT if is_pass else RED_LIGHT

        vals = [idx+1, r["id"], r["name"], r["status"], r.get("note","")]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col)
            cell.value = val
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col in [1,2,4] else "left", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=bg)
            if col == 4:
                cell.font = Font(bold=True, color=GREEN if is_pass else RED)
        ws.row_dimensions[row].height = 18

    wb.save("ket_qua_test.xlsx")
    print("✅ Đã xuất: ket_qua_test.xlsx")

xuat_html()
xuat_excel()
print("\n🎯 Mở file để chụp màn hình:")
print("   HTML : ket_qua_test.html  (mở bằng Chrome)")
print("   Excel: ket_qua_test.xlsx  (mở bằng Excel)")

# Tự mở file
import subprocess
subprocess.Popen(["start", "ket_qua_test.html"], shell=True)
subprocess.Popen(["start", "ket_qua_test.xlsx"], shell=True)
